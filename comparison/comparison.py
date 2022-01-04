import os
from openpyxl import load_workbook

path_as400 = r"D:\comparison\as400file"
path_sit = r"D:\comparison\sitfile"
path_result = r"D:\result.txt"
path_format = r"D:\comparison\報送格式.xlsx"


def listFile(path):
    '''透過指定路徑讀取該路經所有檔案，回傳檔名list'''
    try:
        a = os.listdir(path)
        return(a)
    except:
        input(f"該路徑不存在:==>{path}")


def decoding(path):
    '''轉編碼區塊，若有需要'''
    pass


'''比對檔案主區塊'''


def compareFile(filename):
    as400_file_path = os.path.join(path_as400, filename)
    sit_file_path = os.path.join(path_sit, filename)
    filecode = filename.split(".")[1]
    try:
        wb = load_workbook(filename=path_format)
    except:
        input('查無報送格式，輸入任意鍵退出')

    try:
        sheet = wb[filecode]
    except KeyError:
        return "格式不存在，略過比對"

    '''readlines宣告'''
    as400_readlines = open(as400_file_path, 'r',
                           encoding="Big5", errors="ignore").readlines()
    sit_readlines = open(sit_file_path, 'r', encoding="Big5",
                         errors="ignore").readlines()

    '''頭尾資料長度比較'''
    a_Header = as400_readlines[0]
    s_Header = sit_readlines[0]
    a_Footer = as400_readlines[-1]
    s_Footer = sit_readlines[-1]

    if len(a_Header) == len(s_Header):
        as400_readlines.pop(0)
        sit_readlines.pop(0)
    else:
        return "首筆資料長度不相等，略過比對"

    if a_Header == s_Header:
        pass
    else:
        return "首筆資料不相等，略過比對"

    if len(a_Footer) == len(s_Footer):
        pass
    else:
        return "末筆資料長度不相等，略過比對"
    if a_Footer == s_Footer:
        as400_readlines.pop(-1)
        sit_readlines.pop(-1)
    else:
        return "末筆資料不相等，略過比對"

    '''讀取格式'''
    mr = sheet.max_row

    set_list = {}  # 格式dict {欄位名稱:長度}
    key_list = []  # Key值欄位名稱
    column_list_A = []  # AS400 column值dict
    column_list_S = []  # SIT column值dict
    for i in range(2, mr+1):
        column_name = sheet['A'+str(i)].value
        column_length = sheet['D'+str(i)].value
        if sheet['B'+str(i)].value == "V":
            key_list.append(column_name)
        set_list[column_name] = column_length
    # print(as400_readlines)

    '''拆分各行資料存入json'''
    for al in as400_readlines:
        start = 0
        A_dict = {}
        A_dict = dict(no=as400_readlines.index(al)+1)
        for sing in set_list:
            end = start + set_list[sing]
            A_dict.update([(sing, al[start:end])])
            start = end

        '''以list存入as400拆分後的筆數+各欄位值'''
        column_list_A.append(A_dict)

    for sl in sit_readlines:
        start = 0
        S_dict = {}
        S_dict = dict(no=sit_readlines.index(sl)+1)
        for sing in set_list:
            end = start + set_list[sing]
            S_dict.update([(sing, sl[start:end])])
            start = end

        '''以list存入sit拆分後的筆數+各欄位值'''
        column_list_S.append(S_dict)

    '''檢查總筆數'''
    if len(column_list_A) != len(column_list_S):
        return "資料總筆數不相等，略過比對"

    '''檢核內容'''
    for ca in column_list_A:
        '''檢核key值'''
        ca_str = ""
        check_list_A = []
        for key in key_list:
            ca_str = ca_str + ca[key]

        for sa in column_list_S:
            sa_str = ""
            for key in key_list:
                sa_str = sa_str + sa[key]
            check_list_A.append(sa_str)

        if ca_str in check_list_A:
            '''詳細檢核'''
            a = check_list_A.index(ca_str) + 1
            b = ca["no"]
            for aa in column_list_S:
                if aa["no"] == a:
                    for rr in set_list:
                        if ca[rr] == aa[rr]:
                            pass
                        else:
                            return "第" + str(b) + "資料與比對資料第" + str(a) + "筆相同，但" + rr + "值不同，須檢視"

        else:
            return "第" + str(ca["no"]) + "筆資料無相同key值"


if __name__ == '__main__':
    '''讀取路徑檔名'''
    as400_list = listFile(path_as400)
    sit_list = listFile(path_sit)
    comp_list = []
    '''開啟結果檔'''
    with open(path_result, 'w') as f:
        '''排除缺漏檔案'''
        for al in as400_list:
            for sl in sit_list:
                if al == sl:
                    comp_list.append(al)
        for al in as400_list:
            if al in comp_list:
                pass
            else:
                f.write(f"缺少檔案:{al}，略過比對\r")

        '''比對各檔案'''
        for cl in comp_list:
            result = compareFile(cl)
            f.write(f"檔案:{cl}=>{result}\r")
            # 若檔案皆相同，則result檔不寫入資料，視為成功
            # if result is None:
            #     pass
            # else:
            #     f.write(f"檔案:{al}=>{result}\r")
