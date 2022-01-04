from openpyxl import load_workbook
import os

out_put_path = 'D:/result.txt'


def anounce():
    print('請先將URS確認排程.xlsx放置D槽下，並確認svn已更新')
    with open(out_put_path, 'w') as f:
        pass
    main()


def main():
    input('按下enter繼續執行')
    try:
        wb = load_workbook(r'D:\URS確認排程.xlsx', 'UTF-8')
    except:
        print('無此檔案，請重新檢查')
        main()
    sheet = wb['URS確認']
    trade_list = []
    count = sheet.max_row  # 從2開始算

    print('讀取確認排程交易==================================================')

    for i in range(2,count):
        value = sheet['E'+str(i)].value
        if value is None:
            break
        trade_list.append(value)

    # 檢核是否有重複資料
    set_trade_list = set(trade_list)
    # print(set_trade_list)
    if len(set_trade_list) == len(trade_list):
        print('無重複資料')
    else:
        print('有重複資料')

    print('讀取完成==========================================================')
    matching(trade_list)


def matching(main_trade_list):
    trades = ['L1', 'L2', 'L3', 'L4', 'L5', 'L6', 'L7', 'L8', 'L9']
    # path = 'D:\SVN_SKL\iTX\Project'
    path = 'D:\SKL\iTX\Project'
    count_svn = 0
    for main_trade in trades:
        trade_list = []
        dif_list = []
        print(f'{main_trade}部分開始:')
        use_path = path + '\\' + main_trade
        dirs = os.listdir(use_path)
        # 整理svn程式名稱，移除.java部分
        for trade_name in dirs:
            true_name = trade_name.split('.')[0]
            if true_name[0:2] != main_trade:
                continue
            if true_name[2] == 'R':  # 調Rim程式
                continue
            if len(true_name) != 5:  # 長度不等於五程式
                continue
            trade_list.append(true_name)  # trade_list =>SVN上全部交易

        for single_trade in trade_list:
            if single_trade not in main_trade_list:
                dif_list.append(single_trade)
        makeFile(main_trade, dif_list)
        count_svn = count_svn + len(dif_list)

    print(f'比對完成，比對結果已寫入{out_put_path}')
    print(f'尚有差異:{count_svn}支')
    input('按下enter結束')

def makeFile(a, b):
    with open(out_put_path, 'a') as f:
        f.write(f'{a}有差異如下:{b}\n')
    


if __name__ == '__main__':
    anounce()
