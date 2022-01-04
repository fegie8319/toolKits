from openpyxl import load_workbook
from shutil import copyfile
import os
'''本機路徑路徑宣告'''
java_path = ''
var_path = r"\\192.168.10.8\ifxDoc\ifxfolder\Dev\var\tran"
tim_path = r"\\192.168.10.8\itxDoc\tim"
tom_path = r"\\192.168.10.8\itxDoc\tom"
souce_trade = r'D:\soucetrade.xlsx'

'''svn路徑宣告'''
svn_java = ''
svn_var = r'D:\SKL\iFX\Dev\var\tran'
svn_tim = r'D:\SKL\iTX\tim'
svn_tom = r'D:\SKL\iTX\tom'


def moveVar(tradename):
    trade_class = tradename[0:2]
    trade_path = os.path.join(var_path, trade_class,tradename+".var")
    target_path = os.path.join(svn_var, trade_class,tradename+".var")
    try:
        copyfile(trade_path,target_path)
    except:
        print(f'skip {tradename}.var')
    
def moveTim(tradename):
    trade_class = tradename[0:2]
    trade_path = os.path.join(tim_path, trade_class,tradename+".tim")
    target_path = os.path.join(svn_tim, trade_class,tradename+".tim")
    try:
        copyfile(trade_path,target_path)
    except:
        print(f'skip {tradename}.tim')

def moveTom(tradename):
    trade_class = tradename[0:2]
    trade_path = os.path.join(tom_path, trade_class,tradename+".tom")
    target_path = os.path.join(svn_tom, trade_class,tradename+".tom")
    try:
        copyfile(trade_path,target_path)
    except:
        print(f'skip {tradename}.tom')

def moveTom_OC(tradename):
    trade_class = tradename[0:2]
    trade_path = os.path.join(tom_path, trade_class,tradename+"_OC.tom")
    target_path = os.path.join(svn_tom, trade_class,tradename+"_OC.tom")
    try:
        copyfile(trade_path,target_path)
    except:
        print(f'skip {tradename}_OC.tom')


def main():
    trade_list = []
    '''讀取要搬檔案資料'''
    try:
        wb = load_workbook(filename=souce_trade)
    except:
        print('source trade 資料表不存在!')

    sheet = wb['工作表1']
    rows = sheet.max_row
    for i in range(1, rows+1):
        trade = sheet['A'+str(i)].value
        trade_list.append(trade)
    wb.close()
    print(len(trade_list))
    '''搬檔開始'''
    '''搬java'''
    '''搬var'''
    list(map(moveVar, trade_list))
    '''搬tim'''
    list(map(moveTim, trade_list))
    '''搬tom'''
    list(map(moveTom, trade_list))
    list(map(moveTom_OC,trade_list))

if __name__ == '__main__':
    main()
