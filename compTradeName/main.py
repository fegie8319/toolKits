from openpyxl import load_workbook
import os

result = "D:/tradename result.txt"

urs_position = {}

def getExcelName(urs_path):
    '''取excel交易代號+名稱，return dict'''
    try:
        wb = load_workbook(filename=urs_path)
    except:
        input('查無URS確認檔案，請重新檢查，輸入任意鍵重新執行1')
        main()
    sheet = wb['URS確認']
    trade_list = {}
    count = sheet.max_row  # 從2開始算

    print('讀取確認排程交易==================================================')

    for i in range(2,count):
        '''E:交易代號 F:交易名稱'''
        tradecode = sheet['E'+str(i)].value
        tradename = sheet['F'+str(i)].value
        if tradecode is None:
            break
        trade_list.update({tradecode: tradename.strip()})
        urs_position.update({tradecode:'E'+str(i)})

    print('URS讀取完成')
    wb.close()
    print(count)
    return(trade_list)


def getVarName(trade_path, sub_trade):
    '''取var交易名稱，return dict'''
    print('讀取VAR==================================================')
    var_list = {}
    for sub in sub_trade:
        com_trade_path = trade_path + "\\" + sub
        dirs = os.listdir(com_trade_path)
        for trade in dirs:
            trade_name = trade.split('.')[0]
            with open(com_trade_path+"\\"+trade, 'r', encoding="UTF-8") as f:
                detail = f.readlines()
                var_target = f"[\"[{trade_name}]"
                for read in detail:
                    if read[:9] == var_target:
                        trade_return_name = read[9:-
                                                 4].replace(" ", "").replace("@", "").replace("\t", "")
                        var_list.update({trade_name: trade_return_name})
                        break
    print('VAR讀取完成')
    return(var_list)


def getTxtranCodeName(trancode_path):
    '''取txtrancode交易代號+名稱，return dict'''
    try:
        wb = load_workbook(filename=trancode_path)
    except:
        input('查無tTxtranCode檔案，請重新檢查，輸入任意鍵重新執行')
        main()
    sheet = wb['工作表1']
    trade_list = {}
    count = sheet.max_row
    print('讀取txtrancode交易代號==================================================')

    for i in range(1,count):
        '''A:交易代號 B:交易名稱'''
        tradecode = sheet['A'+str(i)].value
        tradename = sheet['B'+str(i)].value
        if tradecode is None:
            break
        try:
            trade_list.update({tradecode: tradename.strip()})
            print(tradecode+"..."+tradename)
        except:
            print(tradecode+"資料欄有問題，請排除後重新執行")
            input("請按任意鍵退出")

    print('txtrancode交易代號讀取完成')
    wb.close()
    return(trade_list)

def cover(var_name,var_item,urs_path):
    '''改寫確認排程交易名稱'''
    try:
        wb = load_workbook(filename=urs_path)
    except:
        input('查無URS確認檔案，請重新檢查，輸入任意鍵重新執行')
        main()
    sheet = wb['URS確認']
    pos = urs_position[var_name]
    pos_num = pos[1:]
    sheet['F'+pos_num].value = var_item
    # for i in range(2,count):
    #     '''E:交易代號 F:交易名稱'''
    #     tradecode = sheet['E'+str(i)].value
        
    #     if tradecode == var_name:
    #         sheet['F'+str(i)].value = var_item
    #         break
    print(var_item+"已覆蓋")
    wb.save(filename=urs_path)
    wb.close()


def start():
    '''清空結果檔'''
    with open(result, 'w') as f:
        pass

    '''路徑擺放位置宣告'''
    # print("請先update svn var，並將確認排成擺置D槽下，\n若有需要也請更新txtranCode.xlsx，並同樣放置D槽下")
    # input("完成後請按任意鍵繼續執行")
    main()

def main():
    '''路徑'''
    # trade_path = r"D:\SVN_SKL\iFX\Dev\var\Tran"
    trade_path = r"D:\SKL\iFX\Dev\var\tran"
    sub_trade = ["L1", "L2", "L3", "L4", "L5", "L6",
                 "L7", "L8", "L9", "LC", "LD", "LM", "LY"]
    urs_path = r"D:\URS確認排程.xlsx"
    trancode_path = r"D:\txtrancode.xlsx"

    white_list = r"D:\ignore.txt"


    excel_name = {}  # function return value
    var_name = {}
    trancode_name = {}
    count = 0

    '''get every dict'''
    trancode_name = getTxtranCodeName(
        trancode_path)  # 為避免excel匯出資料問題時需修改trancode excel移至優先執行節省時間

    excel_name = getExcelName(urs_path)

    var_name = getVarName(trade_path, sub_trade)

    print("資料讀取完成")

    '''讀取白名單'''
    with open(white_list, 'r') as w:
        ignores = w.readlines()
        for a in ignores:
            wh_key = a.replace("\n", "")
            var_name.pop(wh_key)
    print("白名單讀取完成，開始比對")

    '''開始比對，並寫入資料'''
    with open(result, 'w', encoding="utf-8") as f:
        pass
        for i in var_name.keys():
            var_item = var_name[i]
            '''若var 與excel都有資料'''
            if i in excel_name:
                if excel_name[i] == var_item:
                    if i in trancode_name:
                        if trancode_name[i] == var_item:
                            pass
                        else:
                            f.write(f"{i}:var與txtrancode名稱不相同\n")
                            count = count + 1
                    else:
                        f.write(f"{i}:不存在於txtrancode\n")
                        count = count + 1
                else:
                    f.write(f"{i}:var與excel名稱不相同\n")
                    cover(i,var_item,urs_path)
                    count = count + 1
            else:
                f.write(f"{i}:不存在於確認排程\n")
                count = count + 1
    input(f"所有比對已完成，共有差異:{count}筆，執行結果已寫入tradename result.txt\n請按任意鍵離開:")


if __name__ == '__main__':
    start()
